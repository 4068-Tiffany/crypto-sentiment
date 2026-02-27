import csv
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def export_csv(results, summary, filename=None):
    if not filename:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"sentiment_{ts}.csv"

    with open(filename, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Timestamp", "Source", "Title", "Sentiment", "Score"])
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for r in results:
            writer.writerow([ts, r["source"], r["title"], r["label"], r["compound"]])
        writer.writerow([])
        writer.writerow(["SUMMARY"])
        writer.writerow(["Bullish", summary["counts"]["BULLISH"], str(summary["percentages"]["BULLISH"]) + "%"])
        writer.writerow(["Bearish", summary["counts"]["BEARISH"], str(summary["percentages"]["BEARISH"]) + "%"])
        writer.writerow(["Neutral", summary["counts"]["NEUTRAL"], str(summary["percentages"]["NEUTRAL"]) + "%"])
        writer.writerow(["Overall", summary["overall"], summary["avg_score"]])

    print("  💾 CSV saved: " + filename)
    return filename


def export_excel(results, summary, filename=None):
    if not EXCEL_AVAILABLE:
        print("  ⚠️  openpyxl not installed. Run: pip install openpyxl")
        return

    if not filename:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"sentiment_{ts}.xlsx"

    wb = openpyxl.Workbook()

    # ── Sheet 1: Posts ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Posts"

    headers = ["Timestamp", "Source", "Title", "Sentiment", "Score"]
    header_fill = PatternFill("solid", fgColor="1E1E2E")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    FILLS = {
        "BULLISH": PatternFill("solid", fgColor="D4EDDA"),
        "BEARISH": PatternFill("solid", fgColor="F8D7DA"),
        "NEUTRAL": PatternFill("solid", fgColor="FFF3CD"),
    }

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for row, r in enumerate(results, 2):
        ws.cell(row=row, column=1, value=ts)
        ws.cell(row=row, column=2, value=r["source"])
        ws.cell(row=row, column=3, value=r["title"])
        ws.cell(row=row, column=4, value=r["label"])
        ws.cell(row=row, column=5, value=r["compound"])
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = FILLS[r["label"]]

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10

    # ── Sheet 2: Summary ────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 15

    summary_data = [
        ("Sentiment", "Count", "Percentage"),
        ("Bullish 🟢", summary["counts"]["BULLISH"], str(summary["percentages"]["BULLISH"]) + "%"),
        ("Bearish 🔴", summary["counts"]["BEARISH"], str(summary["percentages"]["BEARISH"]) + "%"),
        ("Neutral ⚪", summary["counts"]["NEUTRAL"], str(summary["percentages"]["NEUTRAL"]) + "%"),
        ("", "", ""),
        ("Overall", summary["overall"], ""),
        ("Avg Score", summary["avg_score"], ""),
        ("Total Posts", summary["total"], ""),
    ]

    for row, data in enumerate(summary_data, 1):
        for col, val in enumerate(data, 1):
            cell = ws2.cell(row=row, column=col, value=val)
            if row == 1:
                cell.fill = header_fill
                cell.font = header_font

    wb.save(filename)
    print("  📊 Excel saved: " + filename)
    return filename


def export_all(results, summary):
    print("\n  💾 Exporting results...")
    export_csv(results, summary)
    export_excel(results, summary)